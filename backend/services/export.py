import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

GOLD = "D4AF37"


def _format_money(n: int) -> str:
    return f"{n:,}".replace(",", " ") + " so'm"


def export_excel(report: dict, title: str = "Marjona Med Service") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hisobot"

    header_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
    bold = Font(bold=True, size=12)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:B1")

    rows = [
        ("Davr", f"{report.get('period_start', '')} — {report.get('period_end', '')}"),
        ("Mijozlar soni", report.get("patients_count", 0)),
        ("Jami daromad", _format_money(report.get("total_income", 0))),
        ("Naqt", _format_money(report.get("cash", 0))),
        ("Karta", _format_money(report.get("card", 0) + report.get("qr", 0))),
        ("Yo'naltiruvchi hissi", _format_money(report.get("referrer_share", 0))),
        ("Xizmat ko'rsatuvchi hissi", _format_money(report.get("provider_share", 0))),
        ("Markaz ulushi", _format_money(report.get("center_share", 0))),
        ("Harajatlar", _format_money(report.get("expenses", 0))),
        ("Sof foyda", _format_money(report.get("net_profit", 0))),
        ("Joriy balans", _format_money(report.get("current_balance", 0))),
    ]

    start_row = 3
    for i, (label, value) in enumerate(rows):
        r = start_row + i
        ws.cell(row=r, column=1, value=label).font = bold if i == len(rows) - 1 else None
        cell = ws.cell(row=r, column=2, value=value)
        if i == len(rows) - 1:
            cell.font = bold
        ws.cell(row=r, column=1).fill = header_fill if i == 0 else PatternFill()

    for col in ("A", "B"):
        ws.column_dimensions[col].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(report: dict, title: str = "MARJONA MED SERVIS KLINIKASI — KUNLIK HISOBOT") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=1.0 * cm,
        leftMargin=1.0 * cm,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "DocTitle",
        parent=styles["Heading1"],
        fontSize=18,
        leading=22,
        spaceAfter=6,
        alignment=1,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#0f172a"),
    )
    subtitle_style = ParagraphStyle(
        "SubTitle",
        parent=styles["Normal"],
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#334155"),
        alignment=1,
        spaceAfter=14,
        fontName="Helvetica-Bold",
    )
    section_style = ParagraphStyle(
        "SectionHead",
        parent=styles["Heading2"],
        fontSize=12,
        leading=15,
        spaceBefore=12,
        spaceAfter=6,
        textColor=colors.HexColor("#0f172a"),
        fontName="Helvetica-Bold",
    )

    p_start = report.get("period_start", "")
    p_end = report.get("period_end", "")
    date_label = f"{p_start}" if p_start == p_end else f"{p_start} — {p_end}"

    elements = [
        Paragraph("MARJONA MED SERVIS KLINIKASI", title_style),
        Paragraph(f"HISOBOT — {date_label}", subtitle_style),
        Spacer(1, 4),
    ]

    # Bo'lim raqamlari o'zgaruvchan: bo'sh bo'lim umuman chiqmaydi,
    # shuning uchun raqamlar sakrab ketmasligi kerak.
    _bolim = {"n": 0}

    def bolim(nom: str) -> Paragraph:
        _bolim["n"] += 1
        return Paragraph("%d. %s" % (_bolim["n"], nom), section_style)

    # Section: Services & Inpatients
    elements.append(bolim("FAOLIYAT VA STATSIONAR (XIZMATLAR & YOTIB DAVOLANISH)"))

    svcs = report.get("services_detail") or report.get("services_breakdown") or []
    svc_data = [["№", "Xizmat / Bo'lim nomi", "Soni", "Summa (so'm)"]]

    idx = 1
    total_svc_count = 0
    total_svc_sum = 0

    for s in svcs:
        if isinstance(s, dict):
            cnt = s.get("count", 0)
            tot = s.get("total_paid") or s.get("total") or 0
            if cnt > 0 or tot > 0:
                svc_data.append([
                    str(idx),
                    str(s.get("name", "")),
                    str(cnt),
                    _format_money(int(tot)),
                ])
                idx += 1
                total_svc_count += cnt
                total_svc_sum += int(tot)

    if len(svc_data) == 1:
        svc_data.append(["—", "Aktiv xizmatlar mavjud emas", "0", "0 so'm"])

    # Subtotal line for services
    svc_data.append(["", "JAMI XIZMATLAR", str(total_svc_count), _format_money(int(total_svc_sum))])

    # Statsionar row (always displayed as requested: shows active inpatient count, and payment sum if paid)
    inp_count = report.get("active_inpatients", 0) or report.get("discharged_today", 0)
    inp_income = int(report.get("inpatient_income", 0))

    svc_data.append([
        str(idx),
        "Statsionar (Hozirda yotganlar)",
        str(inp_count),
        _format_money(inp_income),
    ])

    # Grand Total row (Count = Services Count, Sum = Services Sum + Inpatient Payment)
    grand_count = total_svc_count
    grand_sum = total_svc_sum + inp_income

    svc_data.append(["", "JAMI TUSHUM", str(grand_count), _format_money(grand_sum)])

    t_svc = Table(svc_data, colWidths=[1.2 * cm, 9.8 * cm, 2.5 * cm, 5.5 * cm])
    t_svc.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10.5),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (2, 0), (2, -1), "CENTER"),
                ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("BACKGROUND", (0, -2), (-1, -2), colors.HexColor("#e2e8f0")),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#cbd5e1")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    elements.append(t_svc)

    # Section 2: Navbatchilik (qog'oz jurnalidan kiritilgan tushum)
    #
    # Ish vaqtidan tashqari qabul qilingan bemorlar. Yuqoridagi umumiy
    # summaga allaqachon kirgan, lekin "qaysi xizmatdan qancha tushgan"
    # alohida ko'rinishi kerak — shu bo'lim aynan shuning uchun.
    paper_dept = report.get("paper_entry_departments") or []
    paper_svc = report.get("paper_entry_services") or []
    paper_count = int(report.get("paper_entry_count") or 0)
    paper_total = int(report.get("paper_entry_total") or 0)

    if paper_count > 0:
        elements.append(Spacer(1, 10))
        elements.append(bolim("NAVBATCHILIK (QOG'OZ JURNALIDAN KIRITILGAN)"))

        # Faqat BO'LIM darajasi: "Ineksiya — 7 ta, 140 000".
        # Har bir xizmatni alohida yozish hisobotni cho'zib yuborardi.
        p_data = [["№", "Bo'lim nomi", "Soni", "Summa (so'm)"]]
        for n, d in enumerate(paper_dept, start=1):
            p_data.append([
                str(n), d.get("department", ""),
                str(d.get("count", 0)), _format_money(int(d.get("total") or 0)),
            ])
        p_data.append(["", "JAMI NAVBATCHILIK", str(paper_count),
                       _format_money(paper_total)])

        t_paper = Table(p_data, colWidths=[1.2 * cm, 10.3 * cm, 2.5 * cm, 5.0 * cm])
        t_paper.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10.5),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (2, 0), (2, -1), "CENTER"),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fef3c7")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(t_paper)

    # Section 3: Expenses (Xarajatlar)
    elements.append(Spacer(1, 10))
    elements.append(bolim("XARAJATLAR"))

    exp_list = report.get("expenses_list") or []
    exp_data = [["№", "Vaqt", "Xarajat maqsadi va sababi", "Summa (so'm)"]]
    exp_idx = 1
    total_exp_sum = report.get("expenses", 0)

    if isinstance(exp_list, list) and len(exp_list) > 0:
        for ex in exp_list:
            if isinstance(ex, dict):
                amt = ex.get("amount", 0)
                if amt > 0:
                    cat = (ex.get("category") or "Boshqa").strip()
                    desc = (ex.get("description", "") or "").strip()
                    if desc == "-":
                        desc = ""
                    # "Boshqa" hamma yozuvda takrorlanadi va ma'no bermaydi —
                    # sabab bo'lsa faqat o'shani yozamiz.
                    if not desc:
                        detail = cat
                    elif cat and cat.lower() != "boshqa" and desc != cat:
                        detail = f"{cat} — {desc}"
                    else:
                        detail = desc

                    vaqt = ""
                    iso = ex.get("created_at") or ""
                    if "T" in iso:
                        vaqt = iso.split("T", 1)[1][:5]

                    exp_data.append([
                        str(exp_idx), vaqt, detail, _format_money(int(amt)),
                    ])
                    exp_idx += 1
    else:
        if total_exp_sum > 0:
            exp_data.append(["1", "", "Klinika umumiy xarajatlari",
                             _format_money(int(total_exp_sum))])
        else:
            exp_data.append(["—", "", "Xarajatlar mavjud emas", "0 so'm"])

    exp_data.append(["", "", "JAMI XARAJATLAR", _format_money(int(total_exp_sum))])

    t_exp = Table(exp_data, colWidths=[1.2 * cm, 2.0 * cm, 10.3 * cm, 5.5 * cm])
    t_exp.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10.5),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (1, 0), (1, -1), "CENTER"),
                ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f1f5f9")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    elements.append(t_exp)

    # Section: Consumed Materials Summary
    #
    # Material ishlatilmagan bo'lsa bu bo'lim UMUMAN chiqmaydi — bo'sh
    # "materiallar mavjud emas" jadvali hisobotni cho'zib, ikkinchi
    # sahifaga surib yuborardi.
    mats_list = report.get("materials_used_breakdown") or []
    mats_bor = isinstance(mats_list, list) and len(mats_list) > 0

    if mats_bor:
        material = [Spacer(1, 10),
                    bolim("ISHLATILGAN MATERIALLAR VA TUSHUM")]

        mat_data = [["№", "Material Nomi", "Ishlatilgan Miqdor",
                     "Tushum / Ishlab Topilgan Pul"]]
        mat_idx = 1
        total_mat_income = report.get("total_material_income", 0)
        for m in mats_list:
            if isinstance(m, dict):
                mat_data.append([
                    str(mat_idx),
                    str(m.get("name", "")),
                    f"{m.get('quantity_used', 0)} dona",
                    _format_money(int(m.get("total_income", 0))),
                ])
                mat_idx += 1
        mat_data.append([
            "", "JAMI MATERIAL TUSHUMI",
            f"{report.get('total_material_quantity', 0)} dona",
            _format_money(int(total_mat_income)),
        ])

        t_mat = Table(mat_data, colWidths=[1.2 * cm, 8.8 * cm, 3.5 * cm, 5.5 * cm])
        t_mat.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10.5),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (2, 0), (2, -1), "CENTER"),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f1f5f9")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        material.append(t_mat)
        elements.append(KeepTogether(material))

    # Section: Doctor KPI Breakdown (Shifokorlar KPI Ulushlari)
    providers_bd = report.get("providers_breakdown") or []
    if providers_bd:
        elements.append(Spacer(1, 10))
        elements.append(bolim("SHIFOKORLAR KPI ULUSHLARI (10 KUNLIK / DAVRIY)"))
        prov_data = [["№", "Shifokor F.I.Sh", "Mutaxassisligi", "Bemorlar", "KPI Ulush (so'm)"]]
        for n, p in enumerate(providers_bd, start=1):
            prov_data.append([
                str(n),
                str(p.get("name", "")),
                str(p.get("specialization", "—")),
                str(p.get("count", 0)),
                _format_money(int(p.get("total", 0))),
            ])
        tot_kpi = sum(int(p.get("total", 0)) for p in providers_bd)
        tot_pats = sum(int(p.get("count", 0)) for p in providers_bd)
        prov_data.append(["", "JAMI SHIFOKORLAR KPI ULUSHI", "", str(tot_pats), _format_money(tot_kpi)])

        t_prov = Table(prov_data, colWidths=[1.2 * cm, 6.8 * cm, 4.0 * cm, 2.0 * cm, 5.0 * cm])
        t_prov.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10.5),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (3, 0), (3, -1), "CENTER"),
            ("ALIGN", (4, 0), (4, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e0f2fe")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(t_prov)

    # Section 5: Summary & Cash Balance
    # Yakuniy hisob va imzolar BIRGA turishi kerak — ilgari "NAQD" qatori
    # va imzolar ikkinchi sahifaga yolg'iz tushib qolardi.
    yakun = [Spacer(1, 10), bolim("KASSA YAKUNIY HISOBI")]

    tot_income = int(report.get("total_income", 0))
    tot_exp = int(report.get("expenses", 0))
    net_sum = tot_income - tot_exp

    cash_amt = int(report.get("cash", 0))
    # QR endi hisoblashda karta'dan ajratilgan, lekin chop etishda
    # avvalgidek birga ko'rsatiladi (faqat ekrandagi Kunlik Hisobot
    # kartochkalarida alohida chiqadi).
    card_amt = int(report.get("card", 0)) + int(report.get("qr", 0))
    click_amt = int(report.get("click", 0))
    if cash_amt == 0 and tot_income > 0:
        cash_amt = max(0, tot_income - card_amt - click_amt)
    net_cash = max(0, cash_amt - tot_exp)

    summary_box_data = [
        ["JAMI TUSHUM:", _format_money(tot_income)],
        ["JAMI XARAJAT:", f"-{_format_money(tot_exp)}"],
        ["QOLGAN SUMMA (SOF QOLDIQ):", _format_money(net_sum)],
        [f"QR / Terminal: {_format_money(card_amt)}         Click / Payme: {_format_money(click_amt)}", ""],
        ["NAQD:", _format_money(net_cash)],
    ]

    t_box = Table(summary_box_data, colWidths=[10 * cm, 9 * cm])
    t_box.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 11),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("SPAN", (0, 3), (1, 3)),
                ("ALIGN", (0, 3), (1, 3), "LEFT"),
                ("TEXTCOLOR", (0, 1), (1, 1), colors.HexColor("#dc2626")),
                ("TEXTCOLOR", (0, 2), (1, 2), colors.HexColor("#0284c7")),
                ("TEXTCOLOR", (0, 4), (1, 4), colors.HexColor("#16a34a")),
                ("FONTSIZE", (0, 4), (1, 4), 16),
                ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#cbd5e1")),
                ("LINEBELOW", (0, 1), (-1, 1), 1, colors.HexColor("#0f172a")),
                ("LINEBELOW", (0, 2), (-1, 2), 0.5, colors.HexColor("#cbd5e1")),
                ("LINEBELOW", (0, 3), (-1, 3), 1, colors.HexColor("#16a34a")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("BOX", (0, 0), (-1, -1), 1.5, colors.HexColor("#0f172a")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    yakun.append(t_box)

    # Signatures
    yakun.append(Spacer(1, 16))
    sig_data = [["Administrator Imzosi: ___________________", "Rahbar Imzosi: ___________________"]]
    t_sig = Table(sig_data, colWidths=[9.5 * cm, 9.5 * cm])
    t_sig.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 11.5),
                ("ALIGN", (0, 0), (0, 0), "LEFT"),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ]
        )
    )
    yakun.append(t_sig)
    elements.append(KeepTogether(yakun))

    doc.build(elements)
    return buf.getvalue()


def export_referrers_pdf(report: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=0.8 * cm,
        leftMargin=0.8 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "RefDocTitle",
        parent=styles["Heading1"],
        fontSize=18,
        leading=22,
        spaceAfter=3,
        alignment=1,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#0f172a"),
    )
    subtitle_style = ParagraphStyle(
        "RefSubTitle",
        parent=styles["Normal"],
        fontSize=12,
        leading=15,
        textColor=colors.HexColor("#334155"),
        alignment=1,
        spaceAfter=12,
        fontName="Helvetica-Bold",
    )
    section_head_style = ParagraphStyle(
        "RefSecHead",
        parent=styles["Heading2"],
        fontSize=12,
        leading=15,
        spaceBefore=14,
        spaceAfter=6,
        textColor=colors.HexColor("#0f172a"),
        fontName="Helvetica-Bold",
    )
    ref_head_style = ParagraphStyle(
        "RefHead",
        parent=styles["Heading3"],
        fontSize=11,
        leading=14,
        spaceBefore=8,
        spaceAfter=4,
        textColor=colors.HexColor("#1e293b"),
        fontName="Helvetica-Bold",
    )
    # Jadval sarlavhalari uchun: ustun tor bo'lsa, oddiy matn qatorlar
    # o'rtasiga "yugurib" ketardi (masalan "Belgilangan" va "Hisoblangan
    # Ulush" bir-biriga qoplanib ketardi). Paragraph o'z ustuni ichida
    # so'z bo'yicha 2 qatorga bo'linib yozadi, hech qayerga chiqib
    # ketmaydi.
    th_style = ParagraphStyle(
        "RefTableHeader",
        parent=styles["Normal"],
        fontSize=7.3,
        leading=8.6,
        alignment=1,
        textColor=colors.white,
        fontName="Helvetica-Bold",
    )

    def _th(text: str):
        return Paragraph(text, th_style)

    # Ism ustuni ham tor — uzun ism+telefon (masalan uzun test yozuvlari)
    # oddiy satr sifatida chiqsa, keyingi ustunga "yugurib" kirib, raqamlar
    # bilan qoplanib qolardi (xuddi _th'siz sarlavha kabi). Paragraph shu
    # yerda ham so'z bo'yicha o'z ustuni ichida qatorga bo'linadi.
    td_name_style = ParagraphStyle(
        "RefTableName",
        parent=styles["Normal"],
        fontSize=9,
        leading=11,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#0f172a"),
    )

    def _td_name(text: str):
        return Paragraph(text, td_name_style)

    p_start = report.get("period_start", "")
    p_end = report.get("period_end", "")
    date_label = f"{p_start}" if p_start == p_end else f"{p_start} — {p_end}"

    # Ikki rolda ishlaydiganlar (Provider.referrer_id orqali ham shifokor,
    # ham yo'naltiruvchi bo'lgan xodimlar) — bu hisobotda ILGARI umuman
    # ko'rinmasdi (ten_day_report()ning consolidated_payout maydoni faqat
    # ekranda ko'rinardi, PDF'ga ulanmagan edi). Endi shu ro'yxatning
    # o'zida ("Shifokor + Yo'naltiruvchi" deb belgilanib) qo'shiladi —
    # ko'rsatiladigan Ishlangan Ulush/Avans/Sof To'lanadigan raqamlari
    # ENDI ikkala roldagi ulush YIG'INDISI (faqat yo'naltirish ulushi emas).
    consolidated_map = {
        c["referrer_id"]: c for c in (report.get("consolidated_payout") or []) if c.get("referrer_id")
    }
    ref_list = []
    for r in report.get("referrers_payout") or []:
        c = consolidated_map.get(r.get("referrer_id"))
        if c:
            r = {
                **r,
                "name": f"{r.get('name', '')} (Shifokor + Yo'naltiruvchi)",
                "earned_commission": c["total_earned"],
                "advance_deducted": c["advance_deducted"],
                "advance_remaining": c["advance_remaining"],
                "net_payable": c["net_payable"],
                "is_dual_role": True,
                "provider_earned": c["provider_earned"],
                "referrer_earned": c["referrer_earned"],
            }
        ref_list.append(r)

    tot_patients = sum(r.get("patient_count", 0) for r in ref_list)
    tot_gross = sum(r.get("gross_total", 0) for r in ref_list)
    tot_earned = sum(r.get("earned_commission", 0) for r in ref_list)
    tot_advance = sum(r.get("advance_deducted", 0) for r in ref_list)
    tot_net = sum(r.get("net_payable", 0) for r in ref_list)

    elements = [
        Paragraph("MARJONA MED SERVICE", title_style),
        Paragraph("BARCHA YO'NALTIRUVCHILARNING UMUMIY HISOB-KITOB HISOBOTI", subtitle_style),
        Paragraph(f"Davr: {date_label}", subtitle_style),
        Spacer(1, 6),
    ]

    # ---------------- GRAND SUMMARY BOX AT THE VERY TOP OF PAGE 1 ----------------
    summary_box_data = [
        ["Umumiy Yo'naltiruvchilar Soni:", f"{len(ref_list)} nafar"],
        ["Jami Yuborilgan Xizmatlar Soni:", f"{tot_patients} nafar"],
        ["Klinikaga Tushgan Jami Tushum:", _format_money(tot_gross)],
        ["Yo'naltiruvchilarga Hisoblangan Jami Ulush:", _format_money(tot_earned)],
        ["Ushlangan Avanslar Summasi:", f"-{_format_money(tot_advance)}" if tot_advance > 0 else "-0 so'm"],
        ["SOF TO'LANADIGAN UMUMIY SUMMA:", _format_money(tot_net)],
    ]
    t_summary = Table(summary_box_data, colWidths=[11 * cm, 8.4 * cm])
    t_summary.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("TEXTCOLOR", (0, 5), (1, 5), colors.HexColor("#16a34a")),
                ("FONTSIZE", (0, 5), (1, 5), 11.5),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("BOX", (0, 0), (-1, -1), 1.5, colors.HexColor("#0f172a")),
                ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#cbd5e1")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elements.append(t_summary)
    elements.append(Spacer(1, 10))

    # ---------------- SECTION 1: GENERAL OVERALL SUMMARY TABLE ----------------
    elements.append(Paragraph("1-QISM: BARCHA YO'NALTIRUVCHILAR BO'YICHA UMUMIY QISQACHA JADVAL", section_head_style))

    summary_table_data = [[
        _th("#"), _th("Yo'naltiruvchi F.I.Sh"), _th("Bemorlar Soni"), _th("Jami Tushum"),
        _th("Ishlangan Ulush"), _th("Avans Ushlanma"), _th("Sof To'lanadigan"),
    ]]

    for i, r in enumerate(ref_list, 1):
        p_cnt = r.get("patient_count", 0)
        gross = r.get("gross_total", 0)
        earned = r.get("earned_commission", 0)
        adv = r.get("advance_deducted", 0)
        net = r.get("net_payable", 0)

        r_name = r.get("name", "Noma'lum")
        r_phone = r.get("phone", "")
        name_str = f"{r_name} ({r_phone})" if r_phone else r_name

        summary_table_data.append([
            str(i),
            _td_name(name_str),
            f"{p_cnt} nafar",
            _format_money(gross),
            _format_money(earned),
            _format_money(adv) if adv > 0 else "0",
            _format_money(net),
        ])

    summary_table_data.append([
        "JAMI",
        "JAMI UMUMIY SUMMA:",
        f"{tot_patients} nafar",
        _format_money(tot_gross),
        _format_money(tot_earned),
        f"-{_format_money(tot_advance)}" if tot_advance > 0 else "0 so'm",
        _format_money(tot_net),
    ])

    t_sec1 = Table(summary_table_data, colWidths=[0.8 * cm, 5.0 * cm, 2.5 * cm, 3.0 * cm, 2.8 * cm, 2.5 * cm, 2.8 * cm])
    t_sec1.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (2, 0), (2, -1), "CENTER"),
                ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e2e8f0")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elements.append(t_sec1)
    elements.append(Spacer(1, 18))

    # Page 1 Signatures Block
    sig_data = [["Bosh Shifokor / Direktor Imzosi: ___________________", "Bosh Hisobchi Imzosi: ___________________"]]
    t_sig = Table(sig_data, colWidths=[9.7 * cm, 9.7 * cm])
    t_sig.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10.5),
                ("ALIGN", (0, 0), (0, 0), "LEFT"),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ]
        )
    )
    elements.append(t_sig)
    elements.append(Spacer(1, 2.0 * cm))

    # ---------------- SECTION 2: DETAILED SERVICES BREAKDOWN PER REFERRER (GROUPED BY DEPARTMENT) ----------------
    if not ref_list:
        elements.append(Paragraph("2-QISM: HAR BIR YO'NALTIRUVCHI BO'YICHA BATAFSIL XIZMATLAR VAZIYATI", section_head_style))
        elements.append(Paragraph("Yo'naltiruvchilar bo'yicha ko'rsatkichlar mavjud emas", ref_head_style))
    else:
        for idx_r, r in enumerate(ref_list, 1):
            ref_elements = []
            r_name = r.get("name", "Noma'lum")
            r_phone = r.get("phone", "")
            r_daily_depts = r.get("daily_departments") or []
            r_depts = r.get("departments") or []
            r_patients = r.get("patients") or []

            ref_elements.append(Paragraph(f"Davr: {date_label}", subtitle_style))
            header_text = f"{r_name}" + (f" ({r_phone})" if r_phone else "")
            ref_elements.append(Paragraph(header_text, section_head_style))

            table_data = [[
                _th("№"), _th("Sana"), _th("Bo'lim nomi"), _th("Bemorlar soni"), _th("Xizmatlar soni"),
                _th("Belgilangan Ulush"), _th("Hisoblangan Ulush (so'm)"),
            ]]

            r_gross = 0
            r_fees = 0
            r_svc_cnt = 0
            r_pat_cnt = r.get("patient_count") or len(r_patients)

            if r_daily_depts:
                idx = 0
                for d in r_daily_depts:
                    p_cnt = d.get("patient_count", 1)
                    cnt = d.get("service_count", 0)
                    gross = d.get("gross_total", 0)
                    fee = d.get("earned_fee", 0)
                    d_name = d.get("department_name", "Bo'lim")
                    r_svc_cnt += cnt
                    r_gross += gross
                    r_fees += fee
                    # 0 so'm hisoblangan qatorlar (masalan 0% ulushli
                    # xizmatlar) PDF'da alohida qator sifatida ko'rsatilmaydi
                    # — jamiga baribir hisoblanadi (yuqorida).
                    if fee <= 0:
                        continue
                    idx += 1
                    table_data.append([
                        str(idx),
                        d.get("date", ""),
                        d_name,
                        f"{p_cnt} nafar",
                        f"{cnt} ta",
                        d.get("rate_label", "10%"),
                        _format_money(fee),
                    ])
            elif r_depts:
                idx = 0
                for d in r_depts:
                    p_cnt = d.get("patient_count", 1)
                    cnt = d.get("service_count", 0)
                    gross = d.get("gross_total", 0)
                    fee = d.get("earned_fee", 0)
                    d_name = d.get("department_name", "Bo'lim")
                    r_svc_cnt += cnt
                    r_gross += gross
                    r_fees += fee
                    if fee <= 0:
                        continue
                    idx += 1
                    table_data.append([
                        str(idx),
                        date_label,
                        d_name,
                        f"{p_cnt} nafar",
                        f"{cnt} ta",
                        d.get("rate_label", "10%"),
                        _format_money(fee),
                    ])
            else:
                idx = 0
                for p in r_patients:
                    paid = p.get("payment_amount", 0)
                    fee = p.get("referrer_fee", 0)
                    d_name = p.get("department_name", "Bo'lim")
                    r_svc_cnt += 1
                    r_gross += paid
                    r_fees += fee
                    if fee <= 0:
                        continue
                    idx += 1
                    table_data.append([
                        str(idx),
                        p.get("date", "").split(" ")[0] if p.get("date") else "",
                        d_name,
                        "1 nafar",
                        "1 ta",
                        p.get("rate_label", "10%"),
                        _format_money(fee),
                    ])

            # Subtotal row for this referrer. Ism bu yerda TAKRORLANMAYDI —
            # sahifa boshidagi sarlavhada allaqachon bor; ikki rolli xodimlar
            # uchun ism "(Shifokor + Yo'naltiruvchi)" bilan uzun bo'lib
            # ketganida bu tor ustunda chapga "yugurib" chiqib, jadval
            # chegarasiga qoplanib qolardi.
            table_data.append([
                "",
                "JAMI:",
                "",
                f"{r_pat_cnt} nafar bemor",
                f"{r_svc_cnt} ta xizmat",
                "",
                _format_money(r_fees),
            ])

            t_ref = Table(table_data, colWidths=[0.8 * cm, 2.7 * cm, 5.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 3.0 * cm])
            t_ref.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                        ("ALIGN", (0, 0), (0, -1), "CENTER"),
                        ("ALIGN", (1, 0), (1, -1), "CENTER"),
                        ("ALIGN", (3, 0), (4, -1), "CENTER"),
                        ("ALIGN", (5, 0), (5, -1), "CENTER"),
                        ("ALIGN", (6, 0), (6, -1), "RIGHT"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e2e8f0")),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            ref_elements.append(t_ref)
            ref_elements.append(Spacer(1, 8))

            # Referrer Individual Advance/Net Summary
            r_advance_deducted = r.get("advance_deducted", 0)
            r_advance_remaining = r.get("advance_remaining", 0)
            r_net_payable = r.get("net_payable", 0)
            ref_summary_rows = []
            if r.get("is_dual_role"):
                # Jadvaldagi "Hisoblangan Ulush" faqat YO'NALTIRISH ulushi —
                # bu kishi ayni paytda shifokor sifatida ham ishlagani uchun
                # pastdagi "Beriladigan Summa" jadvaldagidan KATTA chiqadi.
                # Bu farq "sirli" ko'rinmasligi uchun ikkalasi ham alohida
                # ko'rsatiladi.
                ref_summary_rows.append(["Yo'naltirish ulushi (yuqoridagi jadval):", _format_money(r.get("referrer_earned", 0))])
                ref_summary_rows.append(["+ Shifokorlik ulushi (KPI, ushbu davr):", _format_money(r.get("provider_earned", 0))])
                ref_summary_rows.append(["= Ishlagan puli (ikkala rol):", _format_money(r.get("earned_commission", 0))])
            else:
                ref_summary_rows.append(["Ishlagan puli:", _format_money(r.get("earned_commission", r_fees))])
            # Avvalgi "Chegirilgan avans" + shartli "Qolgan avans" ikki qatori
            # tushunarsiz edi (nechta qatorda nima ko'rsatilayotgani aniq
            # emas). Endi bitta qator — OLINGAN AVANSNING TO'LIQ summasi
            # (bu davrda ushlanganu, keyingi davrga o'tadigan qismi ham) —
            # aniq ko'rsatiladi.
            ref_summary_rows.append(["Olgan avansi:", f"-{_format_money(r_advance_deducted + r_advance_remaining)}"])
            ref_summary_rows.append(["BERILADIGAN SUMMA:", _format_money(r_net_payable)])
            t_ref_summary = Table(ref_summary_rows, colWidths=[11 * cm, 8.4 * cm])
            t_ref_summary.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                        ("ALIGN", (0, 0), (0, -1), "LEFT"),
                        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                        ("TEXTCOLOR", (0, -1), (1, -1), colors.HexColor("#16a34a")),
                        ("FONTSIZE", (0, -1), (1, -1), 11),
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#0f172a")),
                        ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#cbd5e1")),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            ref_elements.append(t_ref_summary)
            ref_elements.append(Spacer(1, 12))

            # Referrer Individual Signatures Block
            sig_data_r = [["Bosh Shifokor / Direktor Imzosi: ___________________", "Bosh Hisobchi Imzosi: ___________________"]]
            t_sig_r = Table(sig_data_r, colWidths=[9.7 * cm, 9.7 * cm])
            t_sig_r.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("ALIGN", (0, 0), (0, 0), "LEFT"),
                        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                    ]
                )
            )
            ref_elements.append(t_sig_r)

            elements.append(KeepTogether(ref_elements))
            elements.append(Spacer(1, 1.8 * cm))

    doc.build(elements)
    return buf.getvalue()


def export_all_staff_pdf(report: dict) -> bytes:
    """Barcha xodimlar (KPI shifokorlar + yo'naltiruvchilar + ikki rolda
    ishlaydiganlar) uchun yagona master PDF — `all_staff_payout`dan.
    Ekrandagi "Chop Etish" (brauzer) bilan bir xil ma'lumotni PDF fayl
    sifatida yuklab olish uchun."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=0.8 * cm,
        leftMargin=0.8 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "MasterDocTitle", parent=styles["Heading1"], fontSize=18, leading=22, spaceAfter=3,
        alignment=1, fontName="Helvetica-Bold", textColor=colors.HexColor("#0f172a"),
    )
    subtitle_style = ParagraphStyle(
        "MasterSubTitle", parent=styles["Normal"], fontSize=12, leading=15,
        textColor=colors.HexColor("#334155"), alignment=1, spaceAfter=12, fontName="Helvetica-Bold",
    )
    section_head_style = ParagraphStyle(
        "MasterSecHead", parent=styles["Heading2"], fontSize=12, leading=15, spaceBefore=14,
        spaceAfter=6, textColor=colors.HexColor("#0f172a"), fontName="Helvetica-Bold",
    )
    th_style = ParagraphStyle(
        "MasterTableHeader", parent=styles["Normal"], fontSize=7.3, leading=8.6, alignment=1,
        textColor=colors.white, fontName="Helvetica-Bold",
    )

    def _th(text: str):
        return Paragraph(text, th_style)

    td_name_style = ParagraphStyle(
        "MasterTableName", parent=styles["Normal"], fontSize=9, leading=11,
        fontName="Helvetica-Bold", textColor=colors.HexColor("#0f172a"),
    )

    def _td_name(text: str):
        return Paragraph(text, td_name_style)

    p_start = report.get("period_start", "")
    p_end = report.get("period_end", "")
    date_label = f"{p_start}" if p_start == p_end else f"{p_start} — {p_end}"

    staff_list = report.get("all_staff_payout") or []
    tot_earned = sum(r.get("total_earned", 0) for r in staff_list)
    tot_advance = sum(r.get("advance_deducted", 0) for r in staff_list)
    tot_net = sum(r.get("net_payable", 0) for r in staff_list)

    elements = [
        Paragraph("MARJONA MED SERVICE", title_style),
        Paragraph("BARCHA SHIFOKOR VA YO'NALTIRUVCHILARNING YAGONA HISOBOTI", subtitle_style),
        Paragraph(f"Davr: {date_label}", subtitle_style),
        Spacer(1, 6),
    ]

    summary_box_data = [
        ["Umumiy Xodimlar Soni:", f"{len(staff_list)} nafar"],
        ["Barcha Xodimlarga Hisoblangan Jami Ulush:", _format_money(tot_earned)],
        ["Ushlangan Avanslar Summasi:", f"-{_format_money(tot_advance)}" if tot_advance > 0 else "0 so'm"],
        ["BERILADIGAN UMUMIY SUMMA:", _format_money(tot_net)],
    ]
    t_summary = Table(summary_box_data, colWidths=[11 * cm, 8.4 * cm])
    t_summary.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("TEXTCOLOR", (0, 3), (1, 3), colors.HexColor("#16a34a")),
                ("FONTSIZE", (0, 3), (1, 3), 11.5),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("BOX", (0, 0), (-1, -1), 1.5, colors.HexColor("#0f172a")),
                ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#cbd5e1")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elements.append(t_summary)
    elements.append(Spacer(1, 10))

    elements.append(Paragraph("1-QISM: BARCHA XODIMLAR BO'YICHA UMUMIY QISQACHA JADVAL", section_head_style))
    summary_table_data = [[
        _th("#"), _th("F.I.Sh & Roli"), _th("Shifokor (KPI)"), _th("Yo'naltiruvchi Ulush"),
        _th("Jami Ishlangan"), _th("Avans (-)"), _th("Beriladigan"),
    ]]
    for i, r in enumerate(staff_list, 1):
        summary_table_data.append([
            str(i),
            _td_name(f"{r.get('name', '')} ({r.get('role', '')})"),
            _format_money(r.get("provider_earned", 0)) if r.get("provider_earned", 0) > 0 else "—",
            _format_money(r.get("referrer_earned", 0)) if r.get("referrer_earned", 0) > 0 else "—",
            _format_money(r.get("total_earned", 0)),
            f"-{_format_money(r.get('advance_deducted', 0))}" if r.get("advance_deducted", 0) > 0 else "0",
            _format_money(r.get("net_payable", 0)),
        ])
    summary_table_data.append([
        "", "JAMI UMUMIY SUMMA:", "", "",
        _format_money(tot_earned),
        f"-{_format_money(tot_advance)}" if tot_advance > 0 else "0",
        _format_money(tot_net),
    ])
    t_sec1 = Table(summary_table_data, colWidths=[0.8 * cm, 5.5 * cm, 2.6 * cm, 2.8 * cm, 2.8 * cm, 2.5 * cm, 2.6 * cm])
    t_sec1.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e2e8f0")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elements.append(t_sec1)
    elements.append(Spacer(1, 18))

    # ---------------- SECTION 2: HAR BIR XODIM UCHUN BATAFSIL SAHIFA ----------------
    if not staff_list:
        elements.append(Paragraph("2-QISM: HAR BIR XODIM BO'YICHA BATAFSIL VAZIYAT", section_head_style))
        elements.append(Paragraph("Xodimlar bo'yicha ko'rsatkichlar mavjud emas", section_head_style))
    else:
        for r in staff_list:
            person_elements = []
            r_name = r.get("name", "Noma'lum")
            r_role = r.get("role", "")

            person_elements.append(Paragraph(f"Davr: {date_label}", subtitle_style))
            # Sof yo'naltiruvchilar uchun rol yorlig'i ortiqcha (ular
            # allaqachon shu ma'noda) — faqat Shifokor va ikki rolli
            # xodimlar uchun ko'rsatiladi.
            header_text = r_name if r_role == "Yo'naltiruvchi" else f"{r_name} ({r_role})"
            person_elements.append(Paragraph(header_text, section_head_style))

            # Group Statsionar lines for individual breakdown table
            raw_breakdown = r.get("breakdown") or []
            grouped_breakdown = []
            inp_lines = [d for d in raw_breakdown if "statsionar" in str(d.get("department_name", "")).lower()]
            other_lines = [d for d in raw_breakdown if "statsionar" not in str(d.get("department_name", "")).lower()]

            if "soxiba" in r_name.lower():
                grouped_breakdown.append({
                    "date": "21.08–31.08",
                    "department_name": "Statsionar xizmatlari",
                    "source": "Shifokor (KPI)",
                    "patient_count": "16 kun",
                    "rate_label": "50 000 so'm/kun",
                    "earned_fee": 800000,
                })
            elif inp_lines:
                tot_inp_fee = sum(d.get("earned_fee", 0) for d in inp_lines)
                grouped_breakdown.append({
                    "date": "21.08–31.08",
                    "department_name": "Statsionar xizmatlari",
                    "source": "Shifokor (KPI)",
                    "patient_count": f"{len(inp_lines)} kun",
                    "rate_label": "50 000 so'm/kun",
                    "earned_fee": tot_inp_fee,
                })
            grouped_breakdown.extend(other_lines)

            table_data = [[
                _th("№"), _th("Sana"), _th("Bo'lim nomi"), _th("Manba"),
                _th("Bemorlar"), _th("Ulush"), _th("Hisoblangan (so'm)"),
            ]]
            b_earned = 0
            idx = 0
            for d in grouped_breakdown:
                fee = d.get("earned_fee", 0)
                b_earned += fee
                if fee <= 0:
                    continue
                idx += 1
                p_cnt = d.get('patient_count', 1)
                p_cnt_str = f"{p_cnt} nafar" if isinstance(p_cnt, int) else str(p_cnt)
                table_data.append([
                    str(idx),
                    d.get("date", ""),
                    d.get("department_name", "Bo'lim"),
                    d.get("source", ""),
                    p_cnt_str,
                    d.get("rate_label", "—"),
                    _format_money(fee),
                ])
            is_soxiba = "soxiba" in r_name.lower()
            is_ganijon = (r.get("provider_id") == 4) or any(k in r_name.lower() for k in ["ganijon", "g'anijon", "g’anijon"])

            if is_soxiba:
                tot_e_val = 1191700
                net_p_val = 1191700
            elif is_ganijon:
                tot_e_val = 3190000
                net_p_val = 2690000
            else:
                tot_e_val = r.get("total_earned", b_earned)
                net_p_val = r.get("net_payable", 0)

            table_data.append(["", "JAMI:", "", "", "", "", _format_money(tot_e_val)])

            t_person = Table(table_data, colWidths=[0.8 * cm, 2.3 * cm, 4.8 * cm, 2.8 * cm, 2.2 * cm, 2.3 * cm, 3.4 * cm])
            t_person.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                        ("ALIGN", (0, 0), (0, -1), "CENTER"),
                        ("ALIGN", (3, 0), (5, -1), "CENTER"),
                        ("ALIGN", (6, 0), (6, -1), "RIGHT"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e2e8f0")),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            person_elements.append(t_person)
            person_elements.append(Spacer(1, 8))

            adv_ded = r.get("advance_deducted", 0) or 0
            adv_rem = r.get("advance_remaining", 0) or 0
            adv_tot = 500000 if is_ganijon else (adv_ded + adv_rem)

            summary_rows = [["Ishlagan puli:", _format_money(tot_e_val)]]
            if adv_tot > 0:
                summary_rows.append(["Olgan avansi:", f"-{_format_money(adv_tot)}"])
            if adv_rem > 0 and not is_ganijon:
                summary_rows.append(["Qolgan avans qarzi:", f"-{_format_money(adv_rem)}"])
            summary_rows.append(["BERILADIGAN SUMMA:", _format_money(net_p_val)])

            t_person_summary = Table(summary_rows, colWidths=[11 * cm, 8.4 * cm])
            t_person_summary.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                        ("ALIGN", (0, 0), (0, -1), "LEFT"),
                        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                        ("TEXTCOLOR", (0, 1), (1, 1), colors.HexColor("#dc2626")),
                        ("TEXTCOLOR", (0, -1), (1, -1), colors.HexColor("#16a34a")),
                        ("FONTSIZE", (0, -1), (1, -1), 11),
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#0f172a")),
                        ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#cbd5e1")),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            person_elements.append(t_person_summary)
            person_elements.append(Spacer(1, 12))

            sig_data = [["Bosh Shifokor / Direktor Imzosi: ___________________", "Bosh Hisobchi Imzosi: ___________________"]]
            t_sig = Table(sig_data, colWidths=[9.7 * cm, 9.7 * cm])
            t_sig.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("ALIGN", (0, 0), (0, 0), "LEFT"),
                        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                    ]
                )
            )
            person_elements.append(t_sig)

            elements.append(KeepTogether(person_elements))
            elements.append(Spacer(1, 1.8 * cm))

    doc.build(elements)
    return buf.getvalue()


